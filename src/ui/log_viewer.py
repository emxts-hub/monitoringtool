import os
import json
import csv
import ast
import re
from typing import TYPE_CHECKING
from datetime import datetime, date
from PyQt6.QtCore import Qt, QTimer, QFileSystemWatcher, QObject, QRunnable, QThreadPool, pyqtSignal
from PyQt6.QtGui import QColor, QFont, QCursor, QPainter, QPen
from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QComboBox, 
    QPushButton, QTableWidget, QTableWidgetItem, QHeaderView, 
    QAbstractItemView, QScrollArea, QFrame, QFileDialog, QMessageBox, 
    QApplication, QDialog, QGridLayout
)
from config import get_logs_dir, get_all_logs_dirs

if TYPE_CHECKING:
    from ui.monthly_report import MonthlyReportWidget


class LoadingOverlay(QWidget):
    """Semi-transparent overlay displayed while log operations are in progress."""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setAttribute(Qt.WidgetAttribute.WA_TransparentForMouseEvents, False)
        self.is_dark_theme = True

    def set_theme(self, is_dark_theme):
        self.is_dark_theme = is_dark_theme
        self.update()

    def paintEvent(self, event):
        painter = QPainter(self)
        try:
            painter.setRenderHint(QPainter.RenderHint.Antialiasing)
            mask_color = QColor(13, 17, 23, 180) if self.is_dark_theme else QColor(246, 248, 250, 180)
            painter.fillRect(self.rect(), mask_color)

            card_width, card_height = 200, 60
            cx, cy = self.rect().center().x(), self.rect().center().y()
            card_rect = (cx - card_width // 2, cy - card_height // 2, card_width, card_height)

            bg_color = QColor("#161b22") if self.is_dark_theme else QColor("#ffffff")
            border_color = QColor("#30363d") if self.is_dark_theme else QColor("#d0d7de")
            text_color = QColor("#f0f6fc") if self.is_dark_theme else QColor("#1f2937")

            painter.setPen(QPen(border_color, 1))
            painter.setBrush(bg_color)
            painter.drawRoundedRect(*card_rect, 8, 8)

            painter.setFont(QFont("Segoe UI", 10, QFont.Weight.Bold))
            painter.setPen(text_color)
            painter.drawText(
                cx - card_width // 2,
                cy - card_height // 2,
                card_width,
                card_height,
                Qt.AlignmentFlag.AlignCenter,
                "Loading log data...",
            )
        finally:
            painter.end()


class LogHistoryLoadSignals(QObject):
    finished = pyqtSignal(dict)
    error = pyqtSignal(str)


class LogHistoryLoader(QRunnable):
    def __init__(self, target_dirs, active_lpars, signals):
        super().__init__()
        self.target_dirs = list(target_dirs or [])
        self.active_lpars = sorted({
            LogViewerWidget._normalize_server_name(name)
            for name in active_lpars
            if LogViewerWidget._normalize_server_name(name)
        })
        self.signals = signals

    def run(self):
        try:
            log_data_store = {}
            processed_batches = set()
            discovered_lpars = set(self.active_lpars)

            for target_dir in self.target_dirs:
                if not os.path.exists(target_dir):
                    continue
                files = [
                    f for f in os.listdir(target_dir)
                    if f.endswith(".json") and f.startswith("lpar_history_")
                ]
                files.sort()

                for file_name in files:
                    file_path = os.path.join(target_dir, file_name)
                    try:
                        with open(file_path, "r", encoding="utf-8") as f:
                            data = json.load(f)

                        batches = []
                        if isinstance(data, list):
                            for item in data:
                                if isinstance(item, dict) and "records" in item:
                                    records = item.get("records", [])
                                    if not isinstance(records, list):
                                        records = [records] if isinstance(records, dict) else []
                                    batches.append((item.get("timestamp", ""), records))
                                elif isinstance(item, dict):
                                    batches.append((item.get("timestamp", ""), [item]))
                        elif isinstance(data, dict):
                            records = data.get("records", [])
                            if not isinstance(records, list):
                                records = [records] if isinstance(records, dict) else []
                            batches.append((data.get("timestamp", ""), records))

                        for batch_index, (ts, records) in enumerate(batches):
                            if not ts:
                                continue

                            date_key = ts[:10] if len(ts) >= 10 else date.today().strftime("%Y-%m-%d")
                            if date_key not in log_data_store:
                                log_data_store[date_key] = []

                            for record in records:
                                if isinstance(record, dict):
                                    name = LogViewerWidget._normalize_server_name(
                                        record.get("host_name") or record.get("server_name") or
                                        record.get("lpar") or record.get("server")
                                    )
                                    if name:
                                        discovered_lpars.add(name)

                            record_ids = ",".join(
                                str(record.get("entry_id", ""))
                                for record in records
                                if isinstance(record, dict)
                            )
                            batch_signature = f"{target_dir}_{file_name}_{date_key}_{ts}_{record_ids or batch_index}"
                            if batch_signature not in processed_batches:
                                log_data_store[date_key].append((ts, records))
                                processed_batches.add(batch_signature)
                    except Exception:
                        continue

            self.signals.finished.emit({
                "log_data_store": log_data_store,
                "processed_batches": processed_batches,
                "active_lpars": sorted(discovered_lpars),
            })
        except Exception as err:
            self.signals.error.emit(str(err))


class LogViewerWidget(QWidget):
    monthly_report_widget: "MonthlyReportWidget | None"

    @staticmethod
    def _normalize_server_name(value):
        if value is None:
            return ""
        text = str(value).strip()
        if not text or text.lower() in {"n/a", "none", "unknown", "copy", "copied"}:
            return ""
        if re.fullmatch(r"\d{1,3}(?:\.\d{1,3}){3}", text):
            return ""
        return text

    HOURS = [
        "12AM", "1AM", "2AM", "3AM", "4AM", "5AM", "6AM", "7AM", "8AM", "9AM", "10AM", "11AM",
        "12PM", "1PM", "2PM", "3PM", "4PM", "5PM", "6PM", "7PM", "8PM", "9PM", "10PM", "11PM"
    ]

    def _make_font(self, family="Segoe UI", point_size=9, weight=QFont.Weight.Normal):
        font = QFont(family)
        try:
            size_value = int(point_size)
        except (TypeError, ValueError):
            size_value = 9
        safe_size = max(1, size_value)
        try:
            font.setPointSize(safe_size)
        except Exception:
            font.setPointSize(9)
        font.setWeight(weight)
        return font

    def __init__(self, parent=None):
        super().__init__(parent)
        self.is_dark_theme = True
        self.log_data_store = {}
        self.monthly_report_widget = None
        self._processed_batches = set()
        self.max_history_days = 45
        self.max_batches_per_day = 300
        self.max_total_batches = 1200
        self.active_lpars = []
        self.section_headers = []
        self._history_loading = False
        self._history_reload_pending = False
        self._history_thread_pool = QThreadPool.globalInstance()
        self._last_log_scan_signature = None
        self._last_active_lpars_signature = None
        self._date_selected_by_user = False

        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(16, 16, 16, 16)
        main_layout.setSpacing(12)

        self.scroll_area = QScrollArea()
        scroll = self.scroll_area
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.Shape.NoFrame)
        scroll.setStyleSheet("""
            QScrollArea { background-color: #0d1117; }
            QScrollBar:vertical {
                background: #0d1117;
                width: 10px;
                margin: 0px;
            }
            QScrollBar::handle:vertical {
                background: #30363d;
                min-height: 20px;
                border-radius: 5px;
            }
            QScrollBar::handle:vertical:hover {
                background: #8b949e;
            }
            QScrollBar:horizontal {
                background: #0d1117;
                height: 10px;
                margin: 0px;
            }
            QScrollBar::handle:horizontal {
                background: #30363d;
                min-width: 20px;
                border-radius: 5px;
            }
            QScrollBar::handle:horizontal:hover {
                background: #8b949e;
            }
        """)

        self.container = QWidget()
        container = self.container
        container.setStyleSheet("background-color: #0d1117;")
        self.container_layout = QVBoxLayout(container)
        self.container_layout.setContentsMargins(0, 0, 0, 0)
        self.container_layout.setSpacing(16)

        header_bar = QHBoxLayout()
        self.title_label = QLabel("IBM i LPAR Daily Summary")
        self.title_label.setFont(self._make_font("Segoe UI", 14, QFont.Weight.Bold))
        self.title_label.setStyleSheet("color: #ffffff;")
        header_bar.addWidget(self.title_label)

        header_bar.addStretch()

        self.last_refresh_label = QLabel("Last updated: Never")
        self.last_refresh_label.setFont(self._make_font("Segoe UI", 9, QFont.Weight.Normal))
        self.last_refresh_label.setStyleSheet("color: #8b949e; margin-right: 8px;")
        header_bar.addWidget(self.last_refresh_label)

        date_lbl = QLabel("Select Date:")
        date_lbl.setFont(self._make_font("Segoe UI", 10, QFont.Weight.Bold))
        date_lbl.setStyleSheet("color: #8b949e;")
        header_bar.addWidget(date_lbl)

        self.date_combo = QComboBox()
        self.date_combo.setFixedWidth(130)
        self.date_combo.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.date_combo.currentIndexChanged.connect(self.on_date_changed)
        header_bar.addWidget(self.date_combo)

        btn_export = QPushButton("Export Excel")
        btn_export.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        btn_export.setStyleSheet("""
            QPushButton {
                background-color: #1f6feb;
                color: #ffffff;
                border: 1px solid #388bfd;
                padding: 6px 12px;
                border-radius: 6px;
                font-size: 11px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #388bfd;
            }
        """)
        btn_export.clicked.connect(self.export_to_excel)
        header_bar.addWidget(btn_export)

        self.container_layout.addLayout(header_bar)

        self.container_layout.addWidget(self._create_section_header("ASP Usage"))
        self.asp_table = self._build_matrix_table()
        self.container_layout.addWidget(self.asp_table)

        self.container_layout.addWidget(self._create_section_header("CPU Usage"))
        self.cpu_table = self._build_matrix_table()
        self.container_layout.addWidget(self.cpu_table)

        self.container_layout.addWidget(self._create_section_header("Real-time Refresh Log Stream"))
        self.stream_table = QTableWidget()
        self.stream_table.setColumnCount(8)
        self.stream_table.setHorizontalHeaderLabels([
            "TIMESTAMP", "LPAR NAME", "IP ADDRESS", "CPU USAGE", "ASP USAGE", "SUBSYSTEMS", "STATUS", "SERVICES DOWN"
        ])
        horizontal_header = self.stream_table.horizontalHeader()
        if horizontal_header is not None:
            horizontal_header.setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        vertical_header = self.stream_table.verticalHeader()
        if vertical_header is not None:
            vertical_header.hide()
        self.stream_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self._apply_table_styling(self.stream_table)
        self.stream_table.setMinimumHeight(280)
        self.container_layout.addWidget(self.stream_table)

        scroll.setWidget(container)
        main_layout.addWidget(scroll)

        # Loading Overlay setup
        self.loading_overlay = LoadingOverlay(self)
        self.loading_overlay.hide()

        # File watcher setup
        self.reload_timer = QTimer(self)
        self.reload_timer.setSingleShot(True)
        self.reload_timer.setInterval(1000)
        self.reload_timer.timeout.connect(lambda: self.load_log_history(silent=True))

        self.file_watcher = QFileSystemWatcher(self)
        self.file_watcher.directoryChanged.connect(self._on_logs_changed)
        self.file_watcher.fileChanged.connect(self._on_logs_changed)

        self._setup_file_watcher()

        # 1-Second Auto Refresh Timer (runs silently without popping the overlay)
        self.auto_refresh_timer = QTimer(self)
        self.auto_refresh_timer.setInterval(1000)
        self.auto_refresh_timer.timeout.connect(lambda: self.load_log_history(silent=True))
        self.auto_refresh_timer.start()

        # Initial loading call (shows overlay on app boot)
        QTimer.singleShot(50, lambda: self.load_log_history(silent=False))

    def resizeEvent(self, event):
        super().resizeEvent(event)
        if hasattr(self, "loading_overlay"):
            self.loading_overlay.setGeometry(self.rect())

    def _show_loading(self):
        self.loading_overlay.setGeometry(self.rect())
        self.loading_overlay.show()
        self.loading_overlay.raise_()

    def _hide_loading(self):
        self.loading_overlay.hide()

    def _prune_log_store(self, store):
        if not store:
            return
        while len(store) > self.max_history_days:
            oldest_key = min(store.keys())
            del store[oldest_key]
        for date_key, batches in list(store.items()):
            if len(batches) > self.max_batches_per_day:
                store[date_key] = batches[-self.max_batches_per_day:]
        total_batches = sum(len(batches) for batches in store.values())
        while total_batches > self.max_total_batches:
            oldest_key = min(store.keys())
            if not store.get(oldest_key):
                del store[oldest_key]
                total_batches = sum(len(batches) for batches in store.values())
                continue
            store[oldest_key] = store[oldest_key][1:]
            total_batches = sum(len(batches) for batches in store.values())

    def get_combined_log_data_store(self):
        return self.log_data_store

    def _same_hourly_server_record(self, ts, records, server_name, target_hour):
        if not ts or not server_name:
            return False
        try:
            if ts[:13] != target_hour:
                return False
        except Exception:
            return False
        for rec in records:
            if not isinstance(rec, dict):
                continue
            candidate_names = (
                rec.get("host_name"),
                rec.get("server_name"),
                rec.get("lpar"),
                rec.get("server"),
                rec.get("config_key"),
            )
            for candidate in candidate_names:
                normalized_candidate = self._normalize_server_name(candidate)
                if normalized_candidate and normalized_candidate == self._normalize_server_name(server_name):
                    return True
        return False

    def set_theme(self, is_dark_theme):
        self.is_dark_theme = is_dark_theme
        if is_dark_theme:
            background = "#0d1117"
            text = "#ffffff"
            muted = "#8b949e"
        else:
            background = "#f6f8fa"
            text = "#1f2328"
            muted = "#57606a"

        self.container.setStyleSheet(f"background-color: {background};")
        self.scroll_area.setStyleSheet(f"QScrollArea {{ background-color: {background}; }}")
        self.title_label.setStyleSheet(f"color: {text};")
        self.last_refresh_label.setStyleSheet(f"color: {muted}; margin-right: 8px;")
        for header in self.section_headers:
            header.setStyleSheet(
                f"color: {text}; margin-top: 10px; border: none; background: transparent;"
            )
        self._apply_table_styling(self.asp_table)
        self._apply_table_styling(self.cpu_table)
        self._apply_table_styling(self.stream_table)
        if hasattr(self, "loading_overlay"):
            self.loading_overlay.set_theme(is_dark_theme)
        self.populate_views()

    def _setup_file_watcher(self):
        for logs_dir in get_all_logs_dirs():
            if os.path.exists(logs_dir) and logs_dir not in self.file_watcher.directories():
                self.file_watcher.addPath(logs_dir)

    def _on_logs_changed(self, path):
        self._setup_file_watcher()
        self.reload_timer.start()

    def _update_last_refresh_timestamp(self):
        now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.last_refresh_label.setText(f"Last updated: {now_str}")

    def _create_section_header(self, text):
        lbl = QLabel(text)
        lbl.setFont(self._make_font("Segoe UI", 11, QFont.Weight.Bold))
        header_color = "#ffffff" if self.is_dark_theme else "#1f2328"
        lbl.setStyleSheet(f"color: {header_color}; margin-top: 10px; border: none; background: transparent;")
        self.section_headers.append(lbl)
        return lbl

    def _build_matrix_table(self):
        table = QTableWidget()
        table.setColumnCount(len(self.HOURS) + 2)
        headers = ["LPAR"] + self.HOURS + ["ACTION"]
        table.setHorizontalHeaderLabels(headers)
        vertical_header = table.verticalHeader()
        assert vertical_header is not None
        vertical_header.hide()
        table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        table.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)

        header = table.horizontalHeader()
        assert header is not None
        header.setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(len(self.HOURS) + 1, QHeaderView.ResizeMode.Fixed)
        table.setColumnWidth(len(self.HOURS) + 1, 65)  

        self._apply_table_styling(table)
        return table

    def _apply_table_styling(self, table):
        table.verticalHeader().setDefaultSectionSize(32)
        table.horizontalHeader().setFixedHeight(28)
        if self.is_dark_theme:
            table_background = "#161b22"
            gridline = "#21262d"
            border = "#30363d"
            text = "#c9d1d9"
            header_background = "#0d1117"
            header_text = "#8b949e"
        else:
            table_background = "#ffffff"
            gridline = "#d8dee4"
            border = "#d0d7de"
            text = "#1f2328"
            header_background = "#eaeef2"
            header_text = "#57606a"

        table.setStyleSheet("""
            QTableWidget {
                background-color: %s;
                gridline-color: %s;
                border: 1px solid %s;
                border-radius: 6px;
                color: %s;
            }
            QHeaderView::section {
                background-color: %s;
                color: %s;
                padding: 2px 0px;
                font-weight: bold;
                font-size: 10px;
                border: none;
                border-bottom: 1px solid %s;
            }
            QTableWidget::item {
                border-bottom: 1px solid %s;
                padding: 0px;
                font-size: 11px;
            }
        """ % (table_background, gridline, border, text, header_background, header_text, border, gridline))

    def _update_table_heights(self):
        row_count = len(self.active_lpars)
        header_height = 28
        row_height = 32
        border_padding = 16  
        
        calculated_height = header_height + (row_count * row_height) + border_padding
        dynamic_height = max(80, min(calculated_height, 380))

        self.asp_table.setFixedHeight(dynamic_height)
        self.cpu_table.setFixedHeight(dynamic_height)

    def _on_history_loaded(self, result):
        self._history_loading = False

        loaded_store = result.get("log_data_store", {})
        for d_key, entries in loaded_store.items():
            if d_key not in self.log_data_store:
                self.log_data_store[d_key] = []
            for item in entries:
                if item not in self.log_data_store[d_key]:
                    self.log_data_store[d_key].append(item)

        self._processed_batches.update(result.get("processed_batches", set()))
        normalized_active = [
            self._normalize_server_name(name)
            for name in (self.active_lpars + result.get("active_lpars", []))
        ]
        self.active_lpars = sorted({name for name in normalized_active if name})

        if self.monthly_report_widget is not None:
            self.monthly_report_widget.set_log_data_store(self.log_data_store, source_mode="local")

        current_selection = self.date_combo.currentText()
        today_key = date.today().strftime("%Y-%m-%d")
        self.date_combo.blockSignals(True)
        self.date_combo.clear()

        available_dates = sorted(self.log_data_store.keys(), reverse=True)
        if available_dates:
            self.date_combo.addItems(available_dates)
            if self._date_selected_by_user and current_selection in available_dates:
                self.date_combo.setCurrentText(current_selection)
            elif today_key in available_dates:
                self.date_combo.setCurrentText(today_key)
            else:
                self.date_combo.setCurrentIndex(0)
        else:
            self.date_combo.addItem(date.today().strftime("%Y-%m-%d"))

        self.date_combo.blockSignals(False)
        self.populate_views()
        self._update_last_refresh_timestamp()
        
        # Always guarantee overlay hide when task completes
        self._hide_loading()

        # Handle pending reloads after UI completes update
        if self._history_reload_pending:
            self._history_reload_pending = False
            self.load_log_history(silent=True)

    def _on_history_error(self, err_msg):
        self._history_loading = False
        self._hide_loading()
        print(f"[LogViewerWidget] History load error: {err_msg}")

    def _compute_log_scan_signature(self):
        signature = {}
        for target_dir in get_all_logs_dirs():
            if not os.path.exists(target_dir):
                continue
            files = [
                f for f in os.listdir(target_dir)
                if f.endswith(".json") and f.startswith("lpar_history_")
            ]
            for file_name in sorted(files):
                path = os.path.join(target_dir, file_name)
                try:
                    signature[f"{target_dir}/{file_name}"] = (os.path.getmtime(path), os.path.getsize(path))
                except OSError:
                    continue
        return signature

    def load_log_history(self, active_server_configs=None, silent=False):
        if active_server_configs is not None:
            normalized_names = sorted({
                self._normalize_server_name(name)
                for name in active_server_configs.keys()
                if self._normalize_server_name(name)
            })
            if normalized_names:
                self.active_lpars = normalized_names

        active_signature = tuple(self.active_lpars)
        scan_signature = self._compute_log_scan_signature()
        if self.log_data_store and scan_signature == self._last_log_scan_signature and active_signature == self._last_active_lpars_signature:
            self._update_last_refresh_timestamp()
            return

        if self._history_loading:
            self._history_reload_pending = True
            return

        self._history_loading = True
        self._last_log_scan_signature = scan_signature
        self._last_active_lpars_signature = active_signature
        
        if not silent:
            self._show_loading()

        for target_dir in get_all_logs_dirs():
            if os.path.exists(target_dir):
                for file_name in os.listdir(target_dir):
                    file_path = os.path.join(target_dir, file_name)
                    if file_path not in self.file_watcher.files():
                        try:
                            self.file_watcher.addPath(file_path)
                        except Exception:
                            pass

        self.history_signals = LogHistoryLoadSignals()
        self.history_signals.finished.connect(self._on_history_loaded)
        self.history_signals.error.connect(self._on_history_error)

        loader = LogHistoryLoader(get_all_logs_dirs(), self.active_lpars, self.history_signals)
        if self._history_thread_pool is not None:
            self._history_thread_pool.start(loader)

    def on_date_changed(self):
        if not self.date_combo.signalsBlocked():
            self._date_selected_by_user = True
        self._show_loading()
        self.populate_views()
        QTimer.singleShot(150, self._hide_loading)

    def populate_views(self):
        valid_lpars = [
            name for name in self.active_lpars
            if self._normalize_server_name(name)
        ]
        if valid_lpars:
            self.active_lpars = valid_lpars

        selected_date = self.date_combo.currentText()
        if not selected_date:
            selected_date = date.today().strftime("%Y-%m-%d")

        self.title_label.setText(f"IBM i LPAR Daily Summary ({selected_date})")
        self._update_table_heights()

        combined_store = self.get_combined_log_data_store()
        day_batches = combined_store.get(selected_date, [])

        if not self.active_lpars:
            if self.log_data_store:
                available_dates = sorted(self.log_data_store.keys(), reverse=True)
                if available_dates and self.date_combo.count() == 0:
                    self.date_combo.addItems(available_dates)
                if not self.date_combo.currentText() and available_dates:
                    self.date_combo.setCurrentIndex(0)
            self.asp_table.setRowCount(0)
            self.cpu_table.setRowCount(0)
            self._fill_stream_table(day_batches)
            return

        asp_matrix = {lpar: ["N/A%"] * 24 for lpar in self.active_lpars}
        cpu_matrix = {lpar: ["N/A%"] * 24 for lpar in self.active_lpars}

        seen_usage_slots = set()
        for ts, records in day_batches:
            if not ts or len(ts) < 13:
                continue
            try:
                dt = datetime.strptime(ts[:19], "%Y-%m-%d %H:%M:%S")
                hour_idx = dt.hour
            except ValueError:
                continue

            for rec in records:
                if not isinstance(rec, dict):
                    continue
                lpar = self._normalize_server_name(
                    rec.get("host_name")
                    or rec.get("server_name")
                    or rec.get("lpar")
                    or rec.get("server")
                )
                if not lpar:
                    continue
                usage_key = (lpar, hour_idx)
                if usage_key in seen_usage_slots:
                    continue
                if lpar in asp_matrix:
                    asp_val = rec.get("asp")
                    if isinstance(asp_val, (int, float)):
                        asp_matrix[lpar][hour_idx] = f"{asp_val:.1f}%"
                        seen_usage_slots.add((lpar, hour_idx))

                    cpu_val = rec.get("cpu")
                    if isinstance(cpu_val, (int, float)):
                        cpu_matrix[lpar][hour_idx] = f"{cpu_val:.1f}%"
                        seen_usage_slots.add((lpar, hour_idx))

        self._fill_matrix(self.asp_table, asp_matrix)
        self._fill_matrix(self.cpu_table, cpu_matrix)
        self._fill_stream_table(day_batches)

    def _fill_matrix(self, table, matrix_data):
        table.setRowCount(len(self.active_lpars))
        for row, lpar in enumerate(self.active_lpars):
            lpar_item = QTableWidgetItem(lpar)
            lpar_item.setFont(self._make_font("Segoe UI", 9, QFont.Weight.Bold))
            lpar_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            table.setItem(row, 0, lpar_item)

            row_vals = []
            for col in range(24):
                val = matrix_data[lpar][col]
                row_vals.append(val)
                item = QTableWidgetItem(val)
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                if val != "N/A%":
                    item.setForeground(QColor("#58a6ff"))
                else:
                    item.setForeground(QColor("#484f58"))
                table.setItem(row, col + 1, item)

            copy_btn = QPushButton("Copy")
            copy_btn.setFixedHeight(22)
            copy_btn.setFixedWidth(75)
            copy_btn.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
            copy_btn.setStyleSheet("""
                QPushButton {
                    background-color: #21262d;
                    color: #ffffff;
                    border: 1px solid #363b42;
                    font-size: 11px;
                    font-weight: bold;
                    border-radius: 4px;
                    padding: 0px;
                }
                QPushButton:hover {
                    background-color: #30363d;
                    border-color: #58a6ff;
                    color: #58a6ff;
                }
                QPushButton:pressed {
                    background-color: #161b22;
                }
            """)

            copy_text = "\t".join(row_vals)
            copy_btn.clicked.connect(lambda _, t=copy_text, b=copy_btn: self._copy_to_clipboard(t, b))

            cell_widget = QWidget()
            layout = QHBoxLayout(cell_widget)
            layout.setContentsMargins(2, 2, 2, 2)
            layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
            layout.addWidget(copy_btn)
            table.setCellWidget(row, 25, cell_widget)

    def _copy_to_clipboard(self, text, button):
        clipboard = QApplication.clipboard()
        if clipboard is None:
            return
        clipboard.setText(text)
        button.setText("Copied")
        QTimer.singleShot(1200, lambda: button.setText("Copy"))

    def _fill_stream_table(self, day_batches):
        flat_records = []
        lpar_order = {lpar: i for i, lpar in enumerate(self.active_lpars)}

        for ts, records in reversed(day_batches):
            sorted_batch_records = sorted(
                records,
                key=lambda r: lpar_order.get(
                    self._normalize_server_name(
                        r.get("host_name")
                        or r.get("server_name")
                        or r.get("lpar")
                        or r.get("server")
                    ),
                    999,
                )
            )
            for rec in sorted_batch_records:
                if not isinstance(rec, dict):
                    continue
                lpar = self._normalize_server_name(
                    rec.get("host_name")
                    or rec.get("server_name")
                    or rec.get("lpar")
                    or rec.get("server")
                )
                if not lpar:
                    continue
                if lpar in lpar_order or not lpar_order:
                    flat_records.append((ts, rec))

        self.stream_table.setRowCount(len(flat_records))

        for row, (ts, rec) in enumerate(flat_records):
            lpar = self._normalize_server_name(
                rec.get("host_name")
                or rec.get("server_name")
                or rec.get("lpar")
                or rec.get("server")
            )
            ip = rec.get("ip", "N/A")
            cpu = rec.get("cpu", "N/A")
            asp = rec.get("asp", "N/A")
            status = str(rec.get("status", "Unknown"))
            subs_summary = rec.get("subsystems_summary", "")
            subs_detail = rec.get("subsystems_detail", rec.get("subsystems", []))
            down = rec.get("services_down", rec.get("down_items", "None"))

            self.stream_table.setItem(row, 0, self._table_item(str(rec.get("timestamp", ts)), color="#58a6ff"))
            self.stream_table.setItem(row, 1, self._table_item(lpar, bold=True))
            self.stream_table.setItem(row, 2, self._table_item(ip))

            cpu_str = f"{cpu:.1f}%" if isinstance(cpu, (int, float)) else str(cpu)
            asp_str = f"{asp:.1f}%" if isinstance(asp, (int, float)) else str(asp)
            self.stream_table.setItem(row, 3, self._table_item(cpu_str))
            self.stream_table.setItem(row, 4, self._table_item(asp_str))

            sub_count = f" {subs_summary}" if subs_summary else f"{len(subs_detail)} Active"
            sub_btn = QPushButton(sub_count)
            sub_btn.subs_detail = subs_detail
            sub_btn.setFlat(True)
            sub_btn.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
            sub_btn.setStyleSheet("QPushButton { color: #58a6ff; font-size: 11px; font-weight: bold; border: none; text-decoration: underline; }")
            
            # Disable button if no subsystems to display
            if not subs_detail:
                sub_btn.setEnabled(False)
                sub_btn.setCursor(QCursor(Qt.CursorShape.ArrowCursor))
                sub_btn.setStyleSheet("QPushButton { color: #6e7681; font-size: 11px; font-weight: bold; border: none; }")
            
            expected_key = rec.get("config_key") or lpar
            sub_btn.clicked.connect(
                lambda _, l=lpar, d=subs_detail, k=expected_key:
                self._show_subsystems_dialog(l, d, k)
            )

            sub_container = QWidget()
            sub_layout = QVBoxLayout(sub_container)
            sub_layout.setContentsMargins(0, 0, 0, 0)
            sub_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
            sub_layout.addWidget(sub_btn)
            self.stream_table.setCellWidget(row, 5, sub_container)

            status_badge = QLabel(status)
            status_badge.setAlignment(Qt.AlignmentFlag.AlignCenter)
            status_badge.setFixedHeight(20)
            status_badge.setFixedWidth(90)

            st_upper = status.upper()
            if st_upper in ["NORMAL", "ONLINE"]:
                status_badge.setStyleSheet("background-color: #0d281e; color: #3fb950; border-radius: 10px; font-weight: bold; font-size: 10px;")
            elif st_upper == "ABOVE NORMAL":
                status_badge.setStyleSheet("background-color: #3d2d00; color: #e3b341; border-radius: 10px; font-weight: bold; font-size: 10px;")
            else:
                status_badge.setStyleSheet("background-color: #3c1618; color: #f85149; border-radius: 10px; font-weight: bold; font-size: 10px;")

            badge_container = QWidget()
            badge_layout = QVBoxLayout(badge_container)
            badge_layout.setContentsMargins(0, 0, 0, 0)
            badge_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
            badge_layout.addWidget(status_badge)
            self.stream_table.setCellWidget(row, 6, badge_container)

            down_str = self._format_subsystems_list(down)
            self.stream_table.setItem(row, 7, self._table_item(down_str or "None"))

    def _table_item(self, text, bold=False, color=None):
        if color is None:
            color = "#c9d1d9" if self.is_dark_theme else "#1f2328"
        item = QTableWidgetItem(text)
        item.setTextAlignment(Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignHCenter)
        item.setForeground(QColor(color))
        if bold:
            item.setFont(self._make_font("Segoe UI", 9, QFont.Weight.Bold))
        return item

    def _format_subsystems_list(self, subs_list):
        if not subs_list:
            return "None"
        if isinstance(subs_list, str):
            return subs_list

        formatted_names = []
        for sub in subs_list:
            if isinstance(sub, dict):
                sub_name = sub.get("name") or sub.get("subsystem") or sub.get("service") or str(sub)
                formatted_names.append(str(sub_name))
            elif isinstance(sub, str):
                s_str = sub.strip()
                if s_str.startswith("{") and s_str.endswith("}"):
                    try:
                        parsed = ast.literal_eval(s_str)
                        if isinstance(parsed, dict):
                            sub_name = parsed.get("name") or parsed.get("subsystem") or str(parsed)
                            formatted_names.append(str(sub_name))
                        else:
                            formatted_names.append(s_str)
                    except Exception:
                        formatted_names.append(s_str)
                else:
                    formatted_names.append(s_str)
            else:
                formatted_names.append(str(sub))

        return ", ".join(formatted_names) if formatted_names else "None"

    def export_to_excel(self):
        selected_date = self.date_combo.currentText() or date.today().strftime("%Y-%m-%d")
        default_filename = f"IBM_i_Summary_{selected_date}.xlsx"

        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Export Summary Log",
            default_filename,
            "Excel Workbook (*.xlsx);;CSV File (*.csv)"
        )

        if not file_path:
            return

        try:
            if file_path.endswith(".xlsx"):
                try:
                    import openpyxl
                    wb = openpyxl.Workbook()

                    ws_asp = wb.active
                    if ws_asp is None:
                        raise RuntimeError("Unable to access the active worksheet")
                    ws_asp.title = "ASP"
                    self._export_table_to_openpyxl(self.asp_table, ws_asp)

                    ws_cpu = wb.create_sheet(title="CPU")
                    self._export_table_to_openpyxl(self.cpu_table, ws_cpu)

                    ws_stream = wb.create_sheet(title="Data")
                    self._export_stream_to_openpyxl(ws_stream)

                    wb.save(file_path)
                except ImportError:
                    csv_path = file_path.replace(".xlsx", ".csv")
                    self._export_to_csv(csv_path)
                    QMessageBox.information(
                        self,
                        "Exported as CSV",
                        f"openpyxl module is not installed. Exported log as CSV to:\n{csv_path}"
                    )
                    return
            else:
                self._export_to_csv(file_path)

            QMessageBox.information(self, "Export Successful", f"Log data successfully exported to:\n{file_path}")

        except Exception as e:
            QMessageBox.critical(self, "Export Error", f"An error occurred while exporting:\n{str(e)}")

    def _export_table_to_openpyxl(self, table, sheet):
        headers = [table.horizontalHeaderItem(c).text() for c in range(table.columnCount() - 1)]
        sheet.append(headers)

        for r in range(table.rowCount()):
            row_data = []
            for c in range(table.columnCount() - 1):
                item = table.item(r, c)
                row_data.append(item.text() if item else "")
            sheet.append(row_data)

    def _export_stream_to_openpyxl(self, sheet):
        headers = [
            item.text() if (item := self.stream_table.horizontalHeaderItem(c)) else ""
            for c in range(self.stream_table.columnCount())
        ]
        sheet.append(headers)

        for r in range(self.stream_table.rowCount()):
            row_data = []
            for c in range(self.stream_table.columnCount()):
                if c == 5:
                    widget = self.stream_table.cellWidget(r, c)
                    subs_list = []
                    if widget:
                        btn = widget.findChild(QPushButton)
                        if btn:
                            subs_list = getattr(btn, "subs_detail", [])
                    row_data.append(self._format_subsystems_list(subs_list))
                    continue

                item = self.stream_table.item(r, c)
                if item:
                    row_data.append(item.text())
                else:
                    widget = self.stream_table.cellWidget(r, c)
                    if widget:
                        lbl = widget.findChild(QLabel)
                        row_data.append(lbl.text() if lbl else "")
                    else:
                        row_data.append("")
            sheet.append(row_data)

    def _export_to_csv(self, file_path):
        with open(file_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)

            writer.writerow(["--- ASP USAGE ---"])
            asp_headers = [
                header.text() if header is not None else ""
                for c in range(self.asp_table.columnCount() - 1)
                if (header := self.asp_table.horizontalHeaderItem(c)) is not None
            ]
            writer.writerow(asp_headers)
            for r in range(self.asp_table.rowCount()):
                row_data = []
                for c in range(self.asp_table.columnCount() - 1):
                    item = self.asp_table.item(r, c)
                    row_data.append(item.text() if item is not None else "")
                writer.writerow(row_data)

            writer.writerow([])

            writer.writerow(["--- CPU USAGE ---"])
            cpu_headers = [
                header.text() if header is not None else ""
                for c in range(self.cpu_table.columnCount() - 1)
                if (header := self.cpu_table.horizontalHeaderItem(c)) is not None
            ]
            writer.writerow(cpu_headers)
            for r in range(self.cpu_table.rowCount()):
                row_data = []
                for c in range(self.cpu_table.columnCount() - 1):
                    item = self.cpu_table.item(r, c)
                    row_data.append(item.text() if item is not None else "")
                writer.writerow(row_data)

            writer.writerow([])

            writer.writerow(["--- LOG STREAM ---"])
            stream_headers = []
            for c in range(self.stream_table.columnCount()):
                header = self.stream_table.horizontalHeaderItem(c)
                stream_headers.append(header.text() if header is not None else "")
            writer.writerow(stream_headers)

            for r in range(self.stream_table.rowCount()):
                row_data = []
                for c in range(self.stream_table.columnCount()):
                    if c == 5:
                        widget = self.stream_table.cellWidget(r, c)
                        subs_list = []
                        if widget:
                            btn = widget.findChild(QPushButton)
                            if btn:
                                subs_list = getattr(btn, "subs_detail", [])
                        row_data.append(self._format_subsystems_list(subs_list))
                        continue

                    item = self.stream_table.item(r, c)
                    if item:
                        row_data.append(item.text())
                    else:
                        widget = self.stream_table.cellWidget(r, c)
                        if widget:
                            lbl = widget.findChild(QLabel)
                            row_data.append(lbl.text() if lbl else "")
                        else:
                            row_data.append("")
                writer.writerow(row_data)

    def _show_subsystems_dialog(self, lpar_name, subsystems, expected_key=None):
        from config import EXPECTED_SUBSYSTEMS

        # Don't show dialog if there are no subsystems to display
        if not subsystems:
            return

        dialog = QDialog(self)
        dialog.setWindowTitle(f"Subsystem Status - {lpar_name}")
        dialog.setMinimumWidth(480)
        dialog_bg = "#161b22" if self.is_dark_theme else "#ffffff"
        dialog_text = "#c9d1d9" if self.is_dark_theme else "#1f2328"
        dialog.setStyleSheet(
            f"QDialog {{ background-color: {dialog_bg}; color: {dialog_text}; }}"
        )

        layout = QVBoxLayout(dialog)
        lbl = QLabel(f"Subsystems status on {lpar_name}:")
        lbl.setFont(self._make_font("Segoe UI", 10, QFont.Weight.Bold))
        lbl.setStyleSheet("color: #ffffff; margin-bottom: 8px;")
        layout.addWidget(lbl)

        grid_widget = QWidget()
        grid = QGridLayout(grid_widget)
        grid.setSpacing(6)

        all_display_items = []
        for sub in subsystems:
            sub_name = ""
            status = "ACTIVE"

            if isinstance(sub, dict):
                sub_name = sub.get("name", "")
                status = str(sub.get("status", "ACTIVE")).upper()
            elif isinstance(sub, str):
                s_str = sub.strip()
                if s_str.startswith("{") and s_str.endswith("}"):
                    try:
                        parsed = ast.literal_eval(s_str)
                        if isinstance(parsed, dict):
                            sub_name = parsed.get("name", "")
                            status = str(parsed.get("status", "ACTIVE")).upper()
                    except Exception:
                        sub_name = s_str
                else:
                    sub_name = s_str
            else:
                sub_name = str(sub)

            clean_name = str(sub_name).strip().upper()
            if not clean_name:
                continue
            is_down = status in ["INACTIVE", "DOWN", "INACTIVE/OFF", "OFF"]
            all_display_items.append((clean_name, is_down))

        if not all_display_items:
            return

        for idx, (sub_name, is_down) in enumerate(all_display_items):
            if is_down:
                badge_style = (
                    "background-color: #3c1618; color: #f85149; border: 1px solid #f85149; "
                    "border-radius: 4px; padding: 4px 8px; font-weight: bold; font-size: 11px;"
                )
                badge_text = f"[DOWN] {sub_name}"
            else:
                badge_style = (
                    "background-color: #0d281e; color: #3fb950; border: 1px solid #1e4b33; "
                    "border-radius: 4px; padding: 4px 8px; font-weight: bold; font-size: 11px;"
                )
                badge_text = f"[ACTIVE] {sub_name}"

            badge = QLabel(badge_text)
            badge.setStyleSheet(badge_style)
            grid.addWidget(badge, idx // 3, idx % 3)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.Shape.NoFrame)
        scroll.setWidget(grid_widget)
        layout.addWidget(scroll)

        dialog.exec()

    def closeEvent(self, a0):
        if hasattr(self, "auto_refresh_timer"):
            self.auto_refresh_timer.stop()
        super().closeEvent(a0)

    def get_selected_date(self) -> str:
        return self.date_combo.currentText() or date.today().strftime("%Y-%m-%d")

    def clear_logs(self):
        self.log_data_store.clear()
        self._processed_batches.clear()
        self.date_combo.blockSignals(True)
        self.date_combo.clear()
        self.date_combo.blockSignals(False)
        self.asp_table.setRowCount(0)
        self.cpu_table.setRowCount(0)
        self.stream_table.setRowCount(0)

    def set_dark_theme(self):
        self._apply_table_styling(self.asp_table)
        self._apply_table_styling(self.cpu_table)
        self._apply_table_styling(self.stream_table)