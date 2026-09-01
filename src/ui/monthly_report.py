import csv
from calendar import month_name, monthrange
from collections import defaultdict
from datetime import datetime
import json
import os
import re

from config import get_all_logs_dirs
from PyQt6.QtCore import Qt, QPointF, QThread, pyqtSignal
from PyQt6.QtGui import QColor, QFont, QPainter, QPen, QPolygonF, QCursor
from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QComboBox, QPushButton, QFileDialog, QMessageBox
)


class LoadingOverlay(QWidget):
    """Semi-transparent loading overlay over the main dashboard during background report builds."""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setAttribute(Qt.WidgetAttribute.WA_TransparentForMouseEvents, False)
        self.is_dark_theme = True

    def set_theme(self, is_dark_theme):
        self.is_dark_theme = is_dark_theme
        self.update()

    def paintEvent(self, event):
        painter = QPainter(self)
        try:
            painter.setRenderHint(QPainter.RenderHint.Antialiasing)
            mask_color = QColor(13, 17, 23, 180) if self.is_dark_theme else QColor(246, 248, 250, 180)
            painter.fillRect(self.rect(), mask_color)

            card_width, card_height = 200, 60
            cx, cy = self.rect().center().x(), self.rect().center().y()
            card_rect = (cx - card_width // 2, cy - card_height // 2, card_width, card_height)

            bg_color = QColor("#161b22") if self.is_dark_theme else QColor("#ffffff")
            border_color = QColor("#30363d") if self.is_dark_theme else QColor("#d0d7de")
            text_color = QColor("#f0f6fc") if self.is_dark_theme else QColor("#1f2937")

            painter.setPen(QPen(border_color, 1))
            painter.setBrush(bg_color)
            painter.drawRoundedRect(*card_rect, 8, 8)

            painter.setFont(QFont("Segoe UI", 10, QFont.Weight.Bold))
            painter.setPen(text_color)
            painter.drawText(
                cx - card_width // 2,
                cy - card_height // 2,
                card_width,
                card_height,
                Qt.AlignmentFlag.AlignCenter,
                "Loading report...",
            )
        finally:
            painter.end()


class ReportWorker(QThread):
    """Background worker to calculate report metrics without freezing the UI thread."""
    finished = pyqtSignal(dict, dict)

    def __init__(self, widget, month_key, cpu_mode, asp_mode):
        super().__init__()
        self.widget = widget
        self.month_key = month_key
        self.cpu_mode = cpu_mode
        self.asp_mode = asp_mode

    def run(self):
        cpu_report = self.widget._build_month_report(self.month_key, self.cpu_mode, metric_filter="CPU")
        asp_report = self.widget._build_month_report(self.month_key, self.asp_mode, metric_filter="ASP")
        self.finished.emit(cpu_report, asp_report)


class MonthlyReportWidget(QWidget):
    """Monthly ASP/CPU report styled like a monitoring dashboard, with charts and summary cards."""

    cpu_chart: "MetricChartWidget"
    asp_chart: "MetricChartWidget"

    @staticmethod
    def _normalize_server_name(value):
        if value is None:
            return ""
        text = str(value).strip()
        if not text or text.lower() in {"n/a", "none", "unknown"}:
            return ""
        if re.fullmatch(r"\d{1,3}(?:\.\d{1,3}){3}", text):
            return ""
        return text

    class MetricChartWidget(QWidget):
        def __init__(self, metric_name, parent=None):
            super().__init__(parent)
            self.metric_name = metric_name
            self.mode = "day"
            self.series = []
            self.lines = []
            self.days = []
            self.max_value = 100.0
            self.min_value = 0.0
            self.is_dark_theme = True
            self.base_colors = [
                "#3b82f6", "#f59e0b", "#10b981", "#a78bfa", "#f97316",
                "#34d399", "#f87171", "#60a5fa", "#f9a8d4", "#c084fc"
            ]

        def set_data(self, report_rows, days, current_month):
            normalized_days = []
            for day in days:
                try:
                    normalized_days.append(int(day))
                except (TypeError, ValueError):
                    continue
            if not normalized_days:
                discovered_days = set()
                for row in report_rows:
                    day_map = row.get("day_map", {})
                    if isinstance(day_map, dict):
                        for day in day_map:
                            try:
                                discovered_days.add(int(day))
                            except (TypeError, ValueError):
                                continue
                normalized_days = sorted(discovered_days)

            values_by_server = []
            for row in report_rows:
                server = row.get("server")
                day_map = row.get("day_map", {})
                series = []
                for day in normalized_days:
                    value = day_map.get(day, day_map.get(str(day)))
                    series.append(float(value) if value is not None else None)
                values_by_server.append((server, series, row.get("month_avg", 0.0)))

            self.series = values_by_server
            self.days = normalized_days
            self.current_month = current_month
            self.max_value = max(
                100.0,
                max(
                    (v for _, series, _ in self.series for v in series if v is not None),
                    default=100.0,
                ) * 1.15,
            )
            self.min_value = 0.0
            self.update()

        def set_theme(self, is_dark_theme):
            self.is_dark_theme = is_dark_theme
            self.update()

        def paintEvent(self, a0):
            painter = QPainter(self)
            try:
                painter.setRenderHint(QPainter.RenderHint.Antialiasing)
                background = "#161b22" if self.is_dark_theme else "#f8fafc"
                grid = "#30363d" if self.is_dark_theme else "#cbd5e1"
                muted = "#8b949e" if self.is_dark_theme else "#6b7280"
                legend = "#c9d1d9" if self.is_dark_theme else "#334155"
                painter.fillRect(self.rect(), QColor(background))

                rect = self.rect().adjusted(10, 12, -10, -10)
                left = rect.left() + 26
                right = rect.right() - 8
                top = rect.top() + 18
                bottom = rect.bottom() - 24

                if not self.series or not self.days:
                    painter.setPen(QPen(QColor(muted), 1))
                    painter.drawText(self.rect(), Qt.AlignmentFlag.AlignCenter, "No data available")
                    return

                painter.setPen(QPen(QColor(grid), 1))
                for g_idx in range(0, 6):
                    y = int(top + (bottom - top) * (g_idx / 5.0))
                    painter.drawLine(left, y, right, y)

                x_step = (right - left) / max(1, len(self.days) - 1 if len(self.days) > 1 else 1)
                for idx, day in enumerate(self.days):
                    x = int(round(left + (x_step * idx if len(self.days) > 1 else (right - left) / 2.0)))
                    if idx < len(self.days) - 1:
                        painter.drawLine(x, top, x, bottom)

                for server_index, (name, series, _) in enumerate(self.series):
                    points = []
                    for idx, value in enumerate(series):
                        if value is None:
                            continue
                        if len(self.days) > 1:
                            x = left + (right - left) * idx / (len(self.days) - 1)
                        else:
                            x = left + (right - left) / 2.0
                        y = bottom - ((value - self.min_value) / max(1e-6, self.max_value - self.min_value)) * (bottom - top)
                        points.append(QPointF(x, y))
                    color = QColor(self.base_colors[server_index % len(self.base_colors)])
                    painter.setPen(QPen(color, 2.2))
                    if len(points) == 1:
                        painter.setBrush(color)
                        painter.drawEllipse(points[0], 4, 4)
                        painter.setBrush(Qt.BrushStyle.NoBrush)
                    else:
                        painter.drawPolyline(QPolygonF(points))
                        painter.setBrush(color)
                        for point in points:
                            painter.drawEllipse(point, 3, 3)
                        painter.setBrush(Qt.BrushStyle.NoBrush)

                painter.setPen(QPen(QColor(muted), 1))
                for idx, day in enumerate(self.days):
                    if len(self.days) > 1:
                        x = int(round(left + x_step * idx))
                    else:
                        x = int(round(left + (right - left) / 2.0))
                    if (
                        len(self.days) <= 31
                        or day % max(2, int(len(self.days) / 6)) == 0
                        or idx == len(self.days) - 1
                    ):
                        painter.drawText(int(x) - 10, bottom + 18, 24, 16, Qt.AlignmentFlag.AlignCenter, str(day))

                painter.setPen(QPen(QColor(muted), 1))
                for idx in range(6):
                    value = self.max_value - ((self.max_value - self.min_value) * idx / 5.0)
                    y = top + (bottom - top) * (idx / 5.0)
                    painter.drawText(0, int(y) - 6, 30, 16, Qt.AlignmentFlag.AlignRight, f"{value:.0f}")

                painter.setPen(QPen(QColor("#7c8aa0"), 1))
                painter.drawLine(left, top, left, bottom)
                painter.drawLine(left, bottom, right, bottom)

                start_x = left + 8
                legend_y = top - 8
                for idx, (name, _, _) in enumerate(self.series[:8]):
                    color = QColor(self.base_colors[idx % len(self.base_colors)])
                    painter.setPen(QPen(color, 2.5))
                    painter.drawLine(start_x, legend_y, start_x + 14, legend_y)
                    painter.setPen(QPen(QColor(legend), 1))
                    painter.drawText(start_x + 18, legend_y + 5, 64, 12, Qt.AlignmentFlag.AlignLeft, str(name)[:7])
                    start_x += 80
            finally:
                painter.end()

    class SummaryCard(QWidget):
        def __init__(self, title, value, subtitle, parent=None):
            super().__init__(parent)
            self.setFixedHeight(72)
            self.title = QLabel(title)
            self.title.setFont(QFont("Segoe UI", 9))
            self.title.setStyleSheet("color: #6b7280; background: transparent;")
            self.value = QLabel(value)
            self.value.setFont(QFont("Segoe UI", 12, QFont.Weight.Bold))
            self.value.setStyleSheet("color: #1f2937; background: transparent;")
            self.subtitle = QLabel(subtitle)
            self.subtitle.setFont(QFont("Segoe UI", 8))
            self.subtitle.setStyleSheet("color: #6b7280; background: transparent;")
            layout = QVBoxLayout(self)
            layout.setContentsMargins(12, 10, 12, 10)
            layout.setSpacing(2)
            layout.addWidget(self.title)
            layout.addWidget(self.value)
            layout.addWidget(self.subtitle)

        def set_theme(self, is_dark_theme):
            if is_dark_theme:
                self.setStyleSheet("QWidget { background-color: #161b22; border: 1px solid #30363d; border-radius: 8px; } QLabel { color: #e5e7eb; background: transparent; }")
            else:
                self.setStyleSheet("QWidget { background-color: #ffffff; border: 1px solid #d0d7de; border-radius: 8px; } QLabel { color: #1f2937; background: transparent; }")

    def __init__(self, parent=None):
        super().__init__(parent)
        self.is_dark_theme = True
        self.log_data_store = {}
        self.parent_log_viewer = None
        self.source_mode = "local"
        self.last_sync_at = None
        self._sync_in_progress = False
        self._mode_buttons = []
        self._worker = None

        self.main_layout = QVBoxLayout(self)
        self.main_layout.setContentsMargins(16, 16, 16, 16)
        self.main_layout.setSpacing(12)

        header = QHBoxLayout()
        self.title_label = QLabel("IBM i Monthly ASP/CPU Report")
        self.title_label.setFont(self._make_font("Segoe UI", 14, QFont.Weight.Bold))
        self.title_label.setStyleSheet("color: #ffffff;")
        header.addWidget(self.title_label)
        header.addStretch()

        self.month_combo = QComboBox()
        self.month_combo.setFixedWidth(150)
        self.month_combo.currentIndexChanged.connect(self.refresh_report)
        header.addWidget(self.month_combo)

        self.btn_export = QPushButton("Export to Excel")
        self.btn_export.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.btn_export.clicked.connect(self.export_monthly_to_excel)
        header.addWidget(self.btn_export)

        self.main_layout.addLayout(header)

        self.sync_status_label = QLabel("Source: Local persisted data")
        self.sync_status_label.setFont(self._make_font("Segoe UI", 9, QFont.Weight.Normal))
        self.sync_status_label.setStyleSheet("color: #8b949e;")
        self.main_layout.addWidget(self.sync_status_label)

        self.summary_label = QLabel("Averages are calculated from the 1st to the last day of the selected month.")
        self.summary_label.setFont(self._make_font("Segoe UI", 9, QFont.Weight.Normal))
        self.summary_label.setStyleSheet("color: #8b949e;")
        self.main_layout.addWidget(self.summary_label)

        self.cpu_panel = self._build_metric_section("CPU")
        self.asp_panel = self._build_metric_section("ASP")
        self.main_layout.addWidget(self.cpu_panel)
        self.main_layout.addWidget(self.asp_panel)

        self.loading_overlay = LoadingOverlay(self)
        self.loading_overlay.hide()

        self.set_theme(self.is_dark_theme)
        self.load_month_options()
        self.refresh_report()

    def _extract_hourly_data(self, month_key, metric="cpu"):
        """Builds a structure mapping server -> day_num -> hour_num -> avg metric value."""
        hourly = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
        for rec in self._iter_month_records(month_key):
            ts = str(rec.get("timestamp") or "")
            if len(ts) < 13:
                continue
            try:
                day_num = int(ts[8:10])
                hour_num = int(ts[11:13])
            except ValueError:
                continue
            server = self._normalize_server_name(
                rec.get("host_name") or rec.get("server_name") or rec.get("lpar") or rec.get("server")
            )
            if not server:
                continue
            val = rec.get(metric.lower())
            if isinstance(val, (int, float)) and not isinstance(val, bool):
                hourly[server][day_num][hour_num].append(float(val))

        result = defaultdict(lambda: defaultdict(dict))
        for srv, days in hourly.items():
            for d, hours in days.items():
                for h, vals in hours.items():
                    result[srv][d][h] = round(sum(vals) / len(vals), 2)
        return result

    def export_monthly_to_excel(self):
        """Exports CPU and ASP monthly reports into a summary sheet and separate LPAR tabs with 2-tier heatmap logic."""
        month_key = self.month_combo.currentText() or datetime.now().strftime("%Y-%m")
        default_filename = f"IBM_i_Monthly_Report_{month_key}.xlsx"

        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Export Monthly Report",
            default_filename,
            "Excel Workbook (*.xlsx);;CSV File (*.csv)"
        )

        if not file_path:
            return

        cpu_report = self._build_month_report(month_key, self.cpu_chart.mode, metric_filter="CPU")
        asp_report = self._build_month_report(month_key, self.asp_chart.mode, metric_filter="ASP")

        try:
            if file_path.endswith(".xlsx"):
                try:
                    import openpyxl
                    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
                    from openpyxl.utils import get_column_letter

                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = "Monthly Report"
                    ws.views.sheetView[0].showGridLines = True

                    # Summary Sheet Styles
                    title_font = Font(name="Segoe UI", size=16, bold=True, color="1F2937")
                    sub_font = Font(name="Segoe UI", size=9, italic=True, color="6B7280")
                    section_font = Font(name="Segoe UI", size=12, bold=True, color="1F6FEB")
                    header_font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
                    data_font = Font(name="Segoe UI", size=10, color="1F2937")
                    bold_data_font = Font(name="Segoe UI", size=10, bold=True, color="1F2937")

                    header_fill = PatternFill(start_color="1F6FEB", end_color="1F6FEB", fill_type="solid")
                    zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

                    thin_border = Border(
                        left=Side(style="thin", color="D0D7DE"),
                        right=Side(style="thin", color="D0D7DE"),
                        top=Side(style="thin", color="D0D7DE"),
                        bottom=Side(style="thin", color="D0D7DE")
                    )

                    # Top Summary Header
                    ws["A1"] = f"IBM i Monthly ASP/CPU Report ({month_key})"
                    ws["A1"].font = title_font
                    ws["A2"] = "Source: Local persisted data"
                    ws["A2"].font = sub_font
                    ws["A3"] = "Averages are calculated from the 1st to the last day of the selected month."
                    ws["A3"].font = sub_font

                    current_row = 5

                    def append_metric_table(report, section_title, target_metric):
                        nonlocal current_row
                        ws.cell(row=current_row, column=1, value=section_title).font = section_font
                        current_row += 1

                        days = report.get("days", [])
                        headers = ["Server"] + [f"Day {d}" for d in days] + ["Month Avg"]

                        for col_idx, header in enumerate(headers, start=1):
                            cell = ws.cell(row=current_row, column=col_idx, value=header)
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = Alignment(horizontal="center" if col_idx > 1 else "left", vertical="center")
                            cell.border = thin_border
                        current_row += 1

                        matching_rows = [r for r in report.get("rows", []) if r.get("metric") == target_metric]

                        for r_idx, row in enumerate(matching_rows):
                            day_map = row.get("day_map", {})
                            ws.cell(row=current_row, column=1, value=row.get("server", "")).font = bold_data_font
                            ws.cell(row=current_row, column=1).border = thin_border

                            for d_idx, d in enumerate(days, start=2):
                                val = day_map.get(d, day_map.get(str(d)))
                                cell = ws.cell(row=current_row, column=d_idx)
                                cell.border = thin_border
                                cell.font = data_font
                                cell.alignment = Alignment(horizontal="right")
                                if val is not None:
                                    cell.value = float(val) / 100.0
                                    cell.number_format = "0.00%"

                            avg_cell = ws.cell(row=current_row, column=len(days) + 2)
                            avg_cell.value = float(row.get("month_avg", 0.0)) / 100.0
                            avg_cell.number_format = "0.00%"
                            avg_cell.font = bold_data_font
                            avg_cell.alignment = Alignment(horizontal="right")
                            avg_cell.border = thin_border

                            if r_idx % 2 == 1:
                                for c in range(1, len(days) + 3):
                                    ws.cell(row=current_row, column=c).fill = zebra_fill
                            current_row += 1

                        current_row += 2

                    append_metric_table(cpu_report, "CPU Usage", "CPU")
                    append_metric_table(asp_report, "ASP Usage", "ASP")

                    for col in ws.columns:
                        max_len = max(len(str(cell.value or "")) for cell in col)
                        col_letter = get_column_letter(col[0].column)
                        ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

                    # --- LPAR Individual Sheets Presentation Builder ---
                    year, month_num = map(int, month_key.split("-"))
                    m_name = month_name[month_num]
                    days_in_month = monthrange(year, month_num)[1]

                    hourly_cpu = self._extract_hourly_data(month_key, metric="cpu")
                    hourly_asp = self._extract_hourly_data(month_key, metric="asp")

                    servers = sorted(list(set(
                        [r.get("server") for r in cpu_report.get("rows", [])] +
                        [r.get("server") for r in asp_report.get("rows", [])] +
                        list(hourly_cpu.keys()) +
                        list(hourly_asp.keys())
                    )))

                    # LPAR Sheet Styling & Color Thresholds
                    lpar_title_font = Font(name="Segoe UI", size=14, bold=True, color="1F497D")
                    lpar_section_font = Font(name="Segoe UI", size=12, bold=True, color="1F497D")
                    lpar_header_font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
                    lpar_sub_font = Font(name="Segoe UI", size=9, italic=True, color="595959")
                    day_cell_font = Font(name="Segoe UI", size=10, bold=True, color="1F2937")

                    lpar_header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
                    
                    # 2-Tier Color Rule Fills
                    fill_green = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")   # 0% - 89.99%
                    font_green = Font(name="Segoe UI", size=10, color="276A3C")

                    fill_red = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")     # >= 90%
                    font_red = Font(name="Segoe UI", size=10, bold=True, color="C00000")

                    lpar_border = Border(
                        left=Side(style="thin", color="D9D9D9"),
                        right=Side(style="thin", color="D9D9D9"),
                        top=Side(style="thin", color="D9D9D9"),
                        bottom=Side(style="thin", color="D9D9D9")
                    )

                    for server_name in servers:
                        sheet_name = str(server_name)[:31]
                        lpar_ws = wb.create_sheet(title=sheet_name)
                        lpar_ws.views.sheetView[0].showGridLines = True

                        # Document Header
                        lpar_ws["A1"] = f"LPAR Performance Matrix — {server_name}"
                        lpar_ws["A1"].font = lpar_title_font

                        lpar_ws["A2"] = f"Period: {m_name} ({month_key})"
                        lpar_ws["A2"].font = lpar_sub_font

                        current_r = 4

                        def render_metric_block(section_title, hourly_dict, start_row):
                            r = start_row
                            lpar_ws.cell(row=r, column=1, value=section_title).font = lpar_section_font
                            r += 1

                            # Column Headers
                            day_hdr = lpar_ws.cell(row=r, column=1, value="Day")
                            day_hdr.font = lpar_header_font
                            day_hdr.fill = lpar_header_fill
                            day_hdr.alignment = Alignment(horizontal="center", vertical="center")
                            day_hdr.border = lpar_border

                            for h in range(24):
                                cell = lpar_ws.cell(row=r, column=h + 2, value=f"{h:02d}:00")
                                cell.font = lpar_header_font
                                cell.fill = lpar_header_fill
                                cell.alignment = Alignment(horizontal="center", vertical="center")
                                cell.border = lpar_border

                            lpar_ws.row_dimensions[r].height = 24
                            r += 1

                            # Matrix Rows
                            srv_data = hourly_dict.get(server_name, {})
                            for d in range(1, days_in_month + 1):
                                lpar_ws.row_dimensions[r].height = 20
                                day_cell = lpar_ws.cell(row=r, column=1, value=d)
                                day_cell.font = day_cell_font
                                day_cell.alignment = Alignment(horizontal="center", vertical="center")
                                day_cell.border = lpar_border

                                for h in range(24):
                                    val = srv_data.get(d, {}).get(h)
                                    cell = lpar_ws.cell(row=r, column=h + 2)
                                    cell.border = lpar_border
                                    cell.alignment = Alignment(horizontal="right", vertical="center")

                                    if val is not None:
                                        val_float = float(val)
                                        cell.value = val_float / 100.0
                                        cell.number_format = "0.00%"

                                        # Condition: >= 90% Red | 0–89% Green
                                        if val_float >= 90.0:
                                            cell.fill = fill_red
                                            cell.font = font_red
                                        else:
                                            cell.fill = fill_green
                                            cell.font = font_green
                                    else:
                                        # Empty/Missing Data Cells
                                        cell.value = ""
                                        cell.font = data_font
                                r += 1

                            return r + 2

                        # Build CPU & ASP tables sequentially
                        current_r = render_metric_block("CPU Usage (%)", hourly_cpu, current_r)
                        current_r = render_metric_block("ASP Usage (%)", hourly_asp, current_r)

                        # Set Column Dimensions
                        lpar_ws.column_dimensions["A"].width = 10
                        for h in range(24):
                            col_letter = get_column_letter(h + 2)
                            lpar_ws.column_dimensions[col_letter].width = 11

                    wb.save(file_path)

                except ImportError:
                    csv_path = file_path.rsplit(".", 1)[0] + ".csv"
                    self._write_reports_to_csv(cpu_report, asp_report, csv_path)
                    QMessageBox.information(
                        self,
                        "Exported as CSV",
                        f"openpyxl is not installed. Exported monthly report as CSV instead to:\n{csv_path}"
                    )
                    return
            else:
                self._write_reports_to_csv(cpu_report, asp_report, file_path)

            QMessageBox.information(
                self,
                "Export Successful",
                f"Monthly report successfully exported to:\n{file_path}"
            )

        except Exception as e:
            QMessageBox.critical(
                self,
                "Export Error",
                f"An error occurred while exporting the monthly report:\n{str(e)}"
            )

    def _write_reports_to_csv(self, cpu_report, asp_report, file_path):
        """Fallback writer for CSV outputs."""
        with open(file_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(["IBM i Monthly ASP/CPU Report"])
            writer.writerow(["Source: Local persisted data"])
            writer.writerow([])

            for title, report, target_metric in [("CPU Usage", cpu_report, "CPU"), ("ASP Usage", asp_report, "ASP")]:
                writer.writerow([title])
                days = report.get("days", [])
                writer.writerow(["Server"] + [f"Day {d}" for d in days] + ["Month Avg (%)"])
                matching_rows = [r for r in report.get("rows", []) if r.get("metric") == target_metric]
                for row in matching_rows:
                    day_map = row.get("day_map", {})
                    values = [day_map.get(d, day_map.get(str(d), "")) for d in days]
                    writer.writerow([row.get("server", "")] + values + [row.get("month_avg", 0.0)])
                writer.writerow([])

    def resizeEvent(self, event):
        super().resizeEvent(event)
        self.loading_overlay.setGeometry(self.rect())

    def _build_metric_section(self, metric_name):
        section = QWidget()
        section_layout = QHBoxLayout(section)
        section_layout.setContentsMargins(0, 0, 0, 0)
        section_layout.setSpacing(14)

        left_panel = QWidget()
        left_panel.setMinimumHeight(260)
        left_panel.setStyleSheet("background-color: #f8fafc; border: 1px solid #d0d7de; border-radius: 10px;")
        left_layout = QVBoxLayout(left_panel)
        left_layout.setContentsMargins(10, 10, 10, 8)
        left_layout.setSpacing(6)

        top_row = QHBoxLayout()
        metric_label = QLabel(f"{metric_name} Usage")
        metric_label.setFont(self._make_font("Segoe UI", 11, QFont.Weight.Bold))
        metric_label.setStyleSheet("color: #1f2937;")
        top_row.addWidget(metric_label)
        top_row.addStretch()

        button_group = QWidget()
        button_layout = QHBoxLayout(button_group)
        button_layout.setContentsMargins(0, 0, 0, 0)
        button_layout.setSpacing(3)
        for label in ["Day", "Week", "Month"]:
            btn = QPushButton(label)
            btn.setFixedHeight(24)
            btn.setCheckable(True)
            btn.setAutoExclusive(True)
            if label == "Day":
                btn.setChecked(True)
            btn.setStyleSheet("""
                QPushButton {
                    background-color: transparent;
                    color: #8b949e;
                    border: 1px solid #30363d;
                    border-radius: 10px;
                    padding: 2px 9px;
                }
                QPushButton:hover {
                    background-color: #21262d;
                    color: #f0f6fc;
                }
                QPushButton:checked {
                    background-color: #1f6feb;
                    color: #ffffff;
                    border: 1px solid #388bfd;
                    font-weight: bold;
                }
            """)
            self._mode_buttons.append(btn)
            btn.clicked.connect(lambda _, key=f"{metric_name}:{label}": self._on_toggle_change(key))
            button_layout.addWidget(btn)
        top_row.addWidget(button_group)
        left_layout.addLayout(top_row)

        chart = self.MetricChartWidget(metric_name)
        left_layout.addWidget(chart, stretch=1)
        section_layout.addWidget(left_panel, stretch=3)

        right_panel = QWidget()
        right_panel.setMinimumWidth(230)
        right_panel.setStyleSheet("background-color: transparent;")
        right_panel_layout = QVBoxLayout(right_panel)
        right_panel_layout.setContentsMargins(0, 0, 0, 0)
        right_panel_layout.setSpacing(8)

        card_container = QWidget()
        card_layout = QVBoxLayout(card_container)
        card_layout.setContentsMargins(0, 0, 0, 0)
        card_layout.setSpacing(8)

        self._summary_cards = getattr(self, '_summary_cards', {})
        self._summary_cards[metric_name] = {
            "average": self.SummaryCard("Month Avg", "0.00%", "0 days trend", card_container),
            "servers": QWidget(),
            "chart": chart,
        }

        self._summary_cards[metric_name]["average"].set_theme(self.is_dark_theme)
        card_layout.addWidget(self._summary_cards[metric_name]["average"])

        self._summary_cards[metric_name]["servers"] = QWidget()
        server_layout = QVBoxLayout()
        server_layout.setContentsMargins(0, 0, 0, 0)
        server_layout.setSpacing(6)
        self._replace_layout(
            self._summary_cards[metric_name]["servers"],
            server_layout,
        )
        self._summary_cards[metric_name]["servers"].setStyleSheet("background-color: transparent;")
        card_layout.addWidget(self._summary_cards[metric_name]["servers"])
        right_panel_layout.addWidget(card_container)
        section_layout.addWidget(right_panel, stretch=1)

        setattr(self, f"{metric_name.lower()}_chart", chart)
        setattr(self, f"{metric_name.lower()}_panel", left_panel)
        setattr(self, f"{metric_name.lower()}_label", metric_label)
        setattr(self, f"{metric_name.lower()}_summary_card", self._summary_cards[metric_name]["average"])
        return section

    def _on_toggle_change(self, key):
        metric_name, mode = key.split(":", 1)
        if metric_name == "CPU":
            self.cpu_chart.mode = mode.lower()
        elif metric_name == "ASP":
            self.asp_chart.mode = mode.lower()
        self.refresh_report()

    def _make_font(self, family="Segoe UI", point_size=9, weight=QFont.Weight.Normal):
        font = QFont(family)
        try:
            size_value = int(point_size)
        except (TypeError, ValueError):
            size_value = 9
        font.setPointSize(max(1, size_value))
        font.setWeight(weight)
        return font

    def _clear_layout(self, layout):
        if layout is None:
            return
        while layout.count():
            item = layout.takeAt(0)
            if item is None:
                continue
            widget = item.widget()
            if widget is not None:
                widget.setParent(None)
                widget.deleteLater()
            nested_layout = item.layout()
            if nested_layout is not None:
                self._clear_layout(nested_layout)

    def _replace_layout(self, widget, layout):
        old_layout = widget.layout()
        if old_layout is not None:
            self._clear_layout(old_layout)
            QWidget().setLayout(old_layout)
        widget.setLayout(layout)
        return layout

    def _load_disk_log_store(self):
        merged = {}
        for root_dir in get_all_logs_dirs():
            if not os.path.isdir(root_dir):
                continue
            for current_root, _, files in os.walk(root_dir):
                for file_name in sorted(files):
                    if not file_name.endswith(".json") or not file_name.startswith("lpar_history_"):
                        continue
                    file_path = os.path.join(current_root, file_name)
                    try:
                        with open(file_path, "r", encoding="utf-8") as handle:
                            payload = json.load(handle)
                    except (OSError, ValueError, json.JSONDecodeError):
                        continue

                    batches = []
                    if isinstance(payload, list):
                        for item in payload:
                            if isinstance(item, dict) and "records" in item:
                                records = item.get("records", [])
                                if not isinstance(records, list):
                                    records = [records] if isinstance(records, dict) else []
                                batches.append((item.get("timestamp", ""), records))
                            elif isinstance(item, dict):
                                batches.append((item.get("timestamp", ""), [item]))
                    elif isinstance(payload, dict):
                        records = payload.get("records", [])
                        if not isinstance(records, list):
                            records = [records] if isinstance(records, dict) else []
                        batches.append((payload.get("timestamp", ""), records))

                    for ts, records in batches:
                        if not ts or len(ts) < 10:
                            continue
                        date_key = ts[:10]
                        merged.setdefault(date_key, []).append((ts, records))

        return merged

    def set_log_data_store(self, log_data_store, source_mode=None):
        self.log_data_store = log_data_store if log_data_store is not None else {}
        disk_store = self._load_disk_log_store()
        for date_key, batches in disk_store.items():
            if date_key not in self.log_data_store:
                self.log_data_store[date_key] = []
            for batch in batches:
                if batch not in self.log_data_store[date_key]:
                    self.log_data_store[date_key].append(batch)
        if source_mode is not None:
            self.source_mode = source_mode
        self.load_month_options()
        self.sync_status_label.setText("Source: Local persisted data")
        if self.isVisible():
            self.refresh_report()

    def set_last_sync_timestamp(self, timestamp=None):
        self.last_sync_at = timestamp
        label = getattr(self, "last_sync_label", None)
        if label is not None:
            if timestamp is None:
                text = "Last sync: Never"
            else:
                text = f"Last sync: {timestamp.strftime('%Y-%m-%d %H:%M:%S')}"
            label.setText(text)

    def set_theme(self, is_dark_theme):
        self.is_dark_theme = is_dark_theme
        bg = "#0d1117" if is_dark_theme else "#f6f8fa"
        panel = "#161b22" if is_dark_theme else "#ffffff"
        text = "#c9d1d9" if is_dark_theme else "#1f2328"
        border = "#30363d" if is_dark_theme else "#d0d7de"
        self.setStyleSheet(f"QWidget {{ background-color: {bg}; color: {text}; }} QLabel {{ color: {text}; }} QComboBox {{ background-color: {panel}; color: {text}; border: 1px solid {border}; padding: 4px 8px; }} QPushButton {{ background-color: {panel}; color: {text}; border: 1px solid {border}; border-radius: 4px; }}")
        
        btn_style = "background-color: #1f6feb; color: #ffffff; border: 1px solid #388bfd; font-weight: bold; padding: 4px 10px;"
        if hasattr(self, "btn_export"):
            self.btn_export.setStyleSheet(btn_style)

        self.title_label.setStyleSheet(f"color: {'#ffffff' if is_dark_theme else '#1f2328'};")
        self.sync_status_label.setStyleSheet(f"color: {'#8b949e' if is_dark_theme else '#57606a'};")
        self.summary_label.setStyleSheet(f"color: {'#8b949e' if is_dark_theme else '#57606a'};")
        button_text = "#8b949e" if is_dark_theme else "#57606a"
        button_hover = "#21262d" if is_dark_theme else "#eaeef2"
        button_checked = "#1f6feb"
        for button in self._mode_buttons:
            button.setStyleSheet(
                f"""
                QPushButton {{
                    background-color: transparent;
                    color: {button_text};
                    border: 1px solid {border};
                    border-radius: 10px;
                    padding: 2px 9px;
                }}
                QPushButton:hover {{
                    background-color: {button_hover};
                    color: {text};
                }}
                QPushButton:checked {{
                    background-color: {button_checked};
                    color: #ffffff;
                    border: 1px solid #388bfd;
                    font-weight: bold;
                }}
                """
            )
        for metric_name in ("CPU", "ASP"):
            chart = getattr(self, f"{metric_name.lower()}_chart", None)
            panel_widget = getattr(self, f"{metric_name.lower()}_panel", None)
            label = getattr(self, f"{metric_name.lower()}_label", None)
            if chart is not None:
                chart.set_theme(is_dark_theme)
            if panel_widget is not None:
                panel_widget.setStyleSheet(
                    f"background-color: {panel}; border: 1px solid {border}; border-radius: 10px;"
                )
            if label is not None:
                label.setStyleSheet(f"color: {text};")
        for metric_name in ("CPU", "ASP"):
            section = getattr(self, f"{metric_name.lower()}_summary_card", None)
            if section is not None:
                section.set_theme(is_dark_theme)

        if hasattr(self, "loading_overlay"):
            self.loading_overlay.set_theme(is_dark_theme)

        if self.isVisible():
            self.refresh_report()

    def load_month_options(self):
        disk_store = self._load_disk_log_store()
        merged_store = {**disk_store, **self.log_data_store}
        months = set()
        for date_key in merged_store.keys():
            if isinstance(date_key, str) and len(date_key) >= 7 and date_key[4] == "-":
                months.add(date_key[:7])
        if not months:
            months.add(datetime.now().strftime("%Y-%m"))
        options = sorted(months, reverse=True)
        
        current_selection = self.month_combo.currentText()
        self.month_combo.blockSignals(True)
        self.month_combo.clear()
        self.month_combo.addItems(options)
        if current_selection in options:
            self.month_combo.setCurrentText(current_selection)
        elif self.month_combo.count():
            self.month_combo.setCurrentIndex(0)
        self.month_combo.blockSignals(False)

    def refresh_report(self):
        if not self.isVisible():
            return

        month_key = self.month_combo.currentText() or datetime.now().strftime("%Y-%m")

        self.loading_overlay.setGeometry(self.rect())
        self.loading_overlay.show()
        self.loading_overlay.raise_()

        if self._worker is not None and self._worker.isRunning():
            self._worker.terminate()
            self._worker.wait()

        self._worker = ReportWorker(self, month_key, self.cpu_chart.mode, self.asp_chart.mode)
        self._worker.finished.connect(lambda c_rep, a_rep: self._on_report_ready(c_rep, a_rep, month_key))
        self._worker.start()

    def _on_report_ready(self, cpu_report, asp_report, month_key):
        self._render_metric_panel(self.cpu_chart, self._summary_cards["CPU"], cpu_report, "CPU")
        self._render_metric_panel(self.asp_chart, self._summary_cards["ASP"], asp_report, "ASP")
        self.title_label.setText(f"IBM i Monthly ASP/CPU Report ({month_key})")
        self.loading_overlay.hide()

    def _render_metric_panel(self, chart_widget, summary_group, report, metric_name):
        rows = [row for row in report.get("rows", []) if row.get("metric") == metric_name]
        days = report.get("days", [])
        chart_widget.set_data(rows, days, report.get("month"))

        average_card = summary_group["average"]
        avg_value = round(sum(row.get("month_avg", 0.0) for row in rows) / len(rows), 2) if rows else 0.0
        average_card.value.setText(f"{avg_value:.2f}%")
        average_card.subtitle.setText(f"{len(rows)} server{'s' if len(rows) != 1 else ''} trend")

        server_list = summary_group["servers"]
        server_layout = QVBoxLayout()
        server_layout.setContentsMargins(0, 0, 0, 0)
        server_layout.setSpacing(6)
        self._replace_layout(server_list, server_layout)

        ranking = sorted(rows, key=lambda row: row.get("month_avg", 0.0), reverse=True)
        for index, row in enumerate(ranking):
            label = QLabel(f"{row.get('server', 'Unknown')}")
            label.setFont(self._make_font("Segoe UI", 8))
            label.setStyleSheet(
                f"color: {'#c9d1d9' if self.is_dark_theme else '#475569'}; background: transparent;"
            )
            value = QLabel(f"{row.get('month_avg', 0.0):.2f}%")
            value.setFont(self._make_font("Segoe UI", 8, QFont.Weight.Bold))
            value.setStyleSheet(
                f"color: {'#f0f6fc' if self.is_dark_theme else '#1f2937'}; background: transparent;"
            )
            row_widget = QWidget()
            row_widget.setFixedHeight(26)
            row_layout = QHBoxLayout(row_widget)
            row_layout.setContentsMargins(0, 0, 0, 0)
            row_layout.setSpacing(8)
            dot = QLabel("●")
            color = chart_widget.base_colors[index % len(chart_widget.base_colors)]
            dot.setStyleSheet(
                f"color: {color}; font-size: 9px; background: transparent;"
            )
            row_layout.addWidget(dot)
            row_layout.addWidget(label)
            row_layout.addStretch()
            row_layout.addWidget(value)
            server_layout.addWidget(row_widget)

    def _iter_month_records(self, month_key):
        month_prefix = f"{month_key}-"
        seen = set()

        for date_key, batches in self.log_data_store.items():
            if not isinstance(date_key, str) or not date_key.startswith(month_prefix):
                continue
            if not isinstance(batches, list):
                continue
            for batch in batches:
                if isinstance(batch, tuple) and len(batch) == 2:
                    _, records = batch
                elif isinstance(batch, dict):
                    records = batch.get("records", [])
                else:
                    continue
                if not isinstance(records, list):
                    records = [records] if isinstance(records, dict) else []
                for rec in records:
                    if not isinstance(rec, dict):
                        continue
                    server_id = self._normalize_server_name(
                        rec.get("host_name") or rec.get("server_name") or rec.get("lpar") or rec.get("server")
                    )
                    key = (rec.get("timestamp"), server_id, rec.get("cpu"), rec.get("asp"))
                    if key in seen:
                        continue
                    seen.add(key)
                    yield rec

        for root_dir in get_all_logs_dirs():
            if not os.path.isdir(root_dir):
                continue
            for current_root, _, files in os.walk(root_dir):
                for file_name in sorted(files):
                    if not file_name.startswith("lpar_history_") or not file_name.endswith(".json"):
                        continue
                    if not file_name.startswith(f"lpar_history_{month_key}"):
                        continue

                    date_part = file_name.replace("lpar_history_", "").replace(".json", "")

                    file_path = os.path.join(current_root, file_name)
                    try:
                        with open(file_path, "r", encoding="utf-8") as handle:
                            payload = json.load(handle)
                    except (OSError, ValueError, json.JSONDecodeError):
                        continue

                    batches = payload if isinstance(payload, list) else [payload]
                    for entry in batches:
                        if isinstance(entry, dict) and "records" in entry:
                            records = entry.get("records", [])
                        elif isinstance(entry, dict):
                            records = [entry]
                        else:
                            continue
                        if not isinstance(records, list):
                            records = [records]
                        for rec in records:
                            if not isinstance(rec, dict):
                                continue
                            
                            ts = str(rec.get("timestamp") or entry.get("timestamp") or date_part or "")
                            if not rec.get("timestamp"):
                                rec["timestamp"] = ts

                            server_id = self._normalize_server_name(
                                rec.get("host_name") or rec.get("server_name") or rec.get("lpar") or rec.get("server")
                            )
                            key = (rec.get("timestamp"), server_id, rec.get("cpu"), rec.get("asp"))
                            if key in seen:
                                continue
                            seen.add(key)
                            yield rec

    def _build_month_report(self, month_key, mode="month", metric_filter=None):
        if len(month_key) != 7:
            return {"month": month_key, "days": [], "rows": []}

        try:
            year, month = map(int, month_key.split("-"))
        except ValueError:
            return {"month": month_key, "days": [], "rows": []}

        last_day = monthrange(year, month)[1]
        day_values = defaultdict(lambda: {"cpu": defaultdict(list), "asp": defaultdict(list)})

        for rec in self._iter_month_records(month_key):
            ts = str(rec.get("timestamp") or "")
            if len(ts) < 10:
                continue
            try:
                day_num = int(ts[8:10])
            except ValueError:
                continue
            server = self._normalize_server_name(
                rec.get("host_name") or rec.get("server_name") or rec.get("lpar") or rec.get("server")
            )
            if not server:
                continue
            for metric_name in ("cpu", "asp"):
                value = rec.get(metric_name)
                if isinstance(value, (int, float)) and not isinstance(value, bool):
                    day_values[server][metric_name][day_num].append(float(value))

        metrics_to_process = (
            [("cpu", "CPU")] if metric_filter == "CPU" else 
            [("asp", "ASP")] if metric_filter == "ASP" else 
            [("cpu", "CPU"), ("asp", "ASP")]
        )

        rows = []
        for server in sorted(day_values.keys()):
            for metric_name, metric_label in metrics_to_process:
                daily_map = {}
                for day_num in range(1, last_day + 1):
                    values = day_values.get(server, {}).get(metric_name, {}).get(day_num, [])
                    if values:
                        daily_map[day_num] = round(sum(values) / len(values), 2)
                if not daily_map:
                    continue
                if mode == "week":
                    bucket_map = defaultdict(list)
                    for day_num, value in daily_map.items():
                        bucket_map[(day_num - 1) // 7 + 1].append(value)
                    value_map = {
                        bucket: round(sum(values) / len(values), 2)
                        for bucket, values in bucket_map.items()
                    }
                elif mode == "month":
                    value_map = {1: round(sum(daily_map.values()) / len(daily_map), 2)}
                else:
                    value_map = daily_map
                month_avg = round(sum(value_map.values()) / len(value_map), 2)
                rows.append({
                    "server": server,
                    "metric": metric_label,
                    "day_map": value_map,
                    "month_avg": month_avg,
                })

        if mode == "week":
            periods = list(range(1, (last_day + 6) // 7 + 1))
        elif mode == "month":
            periods = [1]
        else:
            periods = list(range(1, last_day + 1))
        return {"month": month_key, "days": periods, "rows": rows}